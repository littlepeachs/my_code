import tkinter as tk
import openpyxl as pxl
from tkinter import messagebox
from PIL import Image, ImageTk
# from 爬取豆瓣top电影信息 import get_movie_list

gDialog = gSign_in = gWin= leaf_6= None  # 全局变量以g打头，一看就知道是全局变量，全局变量的窗口

ls = []#电影名列表
ls_1=[]#储存用户评分记录的临时列表
dict = {}#单个评分记录的字典
num = 0#记录电影名列表框中选定的电影的序号
count = 0#记录在不关闭登录窗口时是否有多个用户登录，如果是，那么count能保证之前用户的操作仍然有效，其使用在第29行
username_row = -1#默认-1.如果是新用户，就是-1

book_1 = pxl.load_workbook("用户评分记录.xlsx")
sheet_1 = book_1.worksheets[0]
# get_movie_list()#可以使用爬虫爬取250部电影，放在movieTop250.xlsx
book = pxl.load_workbook("movieTop250.xlsx")
sheet = book.worksheets[0]
for cell in sheet['B']:
    ls.append(cell.value)#构建电影列表
del(ls[0])

def btOk_click():#输入用户名界面的确认按钮，要读取该用户名并进行记录
    global username, count,ls_1
    global username_row
    username_row =-1#初始化为-1，如果是新用户，就是-1
    if count>=1:
        sheet_1.append(ls_1)
    ls_1 = []
    username = gDialog.etUsername.get()
    username = username.strip()#去除用户名中的空格
    if (username!=''):
        for i in range(sheet_1.min_row-1,sheet_1.max_row):
            if sheet_1['A'][i].value == username:
                username_row= i
                for cell_1 in sheet_1[i+1]:
                    ls_1.append(cell_1.value)#如果之前已经登录过，就让列表中加入之前的评分记录
                sheet_1.delete_rows(i + 1)#评分记录加入列表后，删除表格中的记录，因为之后还会加回来
                break
        else:
            ls_1.append(username)
        count += 1

        messagebox.showinfo("消息", "您的用户名是：" +
                            username)
        gDialog.destroy()
        show_movie_list()
    else:
        messagebox.showinfo("消息", "用户名不能为空哦")
        gDialog.destroy()
        passwordDialog()

def passwordDialog():#输入用户名界面
    global gDialog

    gDialog = tk.Toplevel(gSign_in)  # 创建对话框窗口
    gDialog.grab_set()  # 显示对话框，并独占输入焦点
    gDialog.title("请输入用户名和密码")
    gDialog.resizable(False, False)
    label1 = tk.Label(gDialog, text="用户名：")
    gDialog.etUsername = tk.Entry(gDialog)#entry构建一个输入框
    label1.grid(row=0, column=0, padx=5, pady=5)
    gDialog.etUsername.grid(row=0, column=1, padx=5, pady=5)

    btOk = tk.Button(gDialog, text="确定", command=btOk_click)
    btOk.grid(row=1, column=0, padx=5, pady=5)
    btCancel = tk.Button(gDialog, text="取消", command=gDialog.destroy)
    btCancel.grid(row=1, column=1, padx=5, pady=5)

def show_movie_list():#电影列表界面，程序的核心窗口
    global gWin
    gWin = tk.Tk()#构建窗口，下同
    gWin.title("Python电影评分系统")#窗口名
    gWin.geometry("520x325")#大小
    gWin.resizable(False, False)  # gWin不可改变大小
    lb_1 = tk.Label(gWin, text="欢迎来到Python电影评分系统", bg="red", fg="white", font=('黑体', 20, 'bold'))#label为一行文字
    lb_1.grid(row=0, column=0, columnspan=4, sticky="EW")#位置，第0行第0列，东西方向对齐
    lb_2 = tk.Label(gWin, text="电影列表", bg="white", fg="black", font=('黑体', 20, 'bold'))
    lb_2.grid(row=1, column=0,columnspan=2,sticky="W")
    btBack = tk.Button(gWin,text="退出",command = gWin.destroy)#构建一个按钮对象
    btBack.grid(row=1, column=2,columnspan=1,sticky="W")
    gWin.nameList = tk.Listbox(gWin,selectmode=tk.SINGLE,exportselection=False)#构建一个列表框对象
    gWin.nameList.grid(row=2, column=0,columnspan=3,padx=3, pady=3,sticky="NSWE")
    for x in ls:
        gWin.nameList.insert(tk.END, x)  # 将电影表格中的电影名插入到列表框尾部
    gWin.nameList.select_set(0, 0)#默认列表框选定的是第0个元素
    scrollbar = tk.Scrollbar(gWin, width=20, orient="vertical",
                             command=gWin.nameList.yview)#构建一个滚动条对象
    gWin.nameList.configure(yscrollcommand=scrollbar.set)  # 绑定listbox和scrollbar
    scrollbar.grid(row=2, column=3, sticky="NS")
    btRank = tk.Button(gWin,text="评分",command = rank)
    btRank.grid(row=3, column=1,sticky="E")
    btIntroduction= tk.Button(gWin, text="使用介绍", command=show_intro)
    btIntroduction.grid(row=3, column=0)
    btInquire = tk.Button(gWin,text="查询",command = check_list)
    btInquire.grid(row=3, column=2)
    gWin.columnconfigure(0, weight=1)
    gWin.rowconfigure(2, weight=1)
    gWin.mainloop()

def record_grade():#记录用户对电影的评分，把评分作为字典放在一个临时列表里
    global dict
    try:#如果用户输入的不是0-10，要对用户进行提醒
        if 0 <= eval(root2.grade.get()) <= 10:
            dict[sheet['B'][num + 1].value] = float(root2.grade.get())
            ls_1.append(str(dict))
            dict.clear()
            for i in range(len(ls_1) - 1):
                if ls_1[i] == None:
                    del (ls_1[i])
            root2.destroy()
            show_movie_list()
        else:
            messagebox.showinfo("消息", "{},一定要在0-10之内评分哦！".format(username))
    except:
        messagebox.showinfo("消息", "{},一定要在0-10之内评分哦！".format(username))

def bt_click():#在核心界面，即电影列表界面，记录鼠标选中的电影，返回该电影在窗口列表框的位置
    sel = gWin.nameList.curselection()#sel形如 (1,2,3)元组类型
    if sel == ():
        gWin.lbHint["text"] = "您还没有选中"
        gWin.lbHint["fg"] = "red"
    else:
        global num
        num = sel[0]
    return num

def rank():#评分
    global num
    num = bt_click()
    gWin.destroy()
    global root2
    root2 = tk.Tk()
    root2.title("Python电影")
    root2.geometry("520x325")
    root2.resizable(False, False)
    lb_1 = tk.Label(root2, text="{}".format(sheet['B'][num + 1].value), bg="red", fg="white", font=('黑体', 20, 'bold'))#电影名
    lb_1.grid(row=0, column=0, columnspan=2, sticky="EW")
    root2.grade = tk.Entry(root2)
    root2.grade.grid(row=1, column=0, padx=5, pady=5)
    btOk = tk.Button(root2, text="确定",command = record_grade)
    btOk.grid(row=1, column=1, padx=5, pady=5)
    lb_2 = tk.Label(root2, text="请输入一个0-10的数", bg="white", fg="black", font=('黑体', 8, 'bold'))
    lb_2.grid(row=2, column=0,columnspan=2)
    btBack = tk.Button(root2, text="返回", command=return_show_list_1)
    btBack.grid(row=2, column=1, padx=5, pady=5)
    root2.columnconfigure(0, weight=1)

def check_list():#查询界面，显示查询的六大按键
    global num
    num = bt_click()
    gWin.destroy()
    global root1
    root1=tk.Tk()
    root1.title("Python电影")
    root1.geometry("500x335")
    root1.resizable(False, False)
    lb_1 = tk.Label(root1, text="{}".format(sheet['B'][num+1].value), bg="red", fg="white", font=('黑体', 20, 'bold'))
    lb_1.grid(row=0, column=0, columnspan=3, sticky="EW")
    # lb_2 = tk.Label(gWin, text="电影列表", bg="white", fg="black", font=('黑体', 20, 'bold'))
    # lb_2.grid(row=1, column=0,sticky="W")
    btkind = tk.Button(root1,text="类型",height=4,width=20,command = show_kind)
    btkind.grid(row=1, column=0,sticky="W")
    btkind.place()
    btplace = tk.Button(root1,text="产地",height=3,width=20,command = show_place)
    btplace.grid(row=1, column=1)
    btall = tk.Button(root1,text="全部信息", height=3, width=20,command=show_all)
    btall.grid(row=2, column=0, sticky="W")
    btrecord = tk.Button(root1,text="用户评分记录", height=3, width=20,command = grade_record)
    btrecord.grid(row=2, column=1)
    btyear = tk.Button(root1,text="年份",height=3,width=20,command = show_year)
    btyear.grid(row=3, column=0,sticky="W")
    btactor = tk.Button(root1,text="演员",height=3,width=20,command = show_people)
    btactor.grid(row=3, column=1)
    bt_return = tk.Button(root1,text="返回",height=2,width=10,command = return_show_list)
    bt_return.grid(row=4, column=0,columnspan = 2)
    root1.columnconfigure(0, weight=1)
    root1.mainloop()

def show_kind():#显示电影的类型
    global leaf_1
    leaf_1 = tk.Tk()
    leaf_1.title("Python电影")
    leaf_1.geometry("520x325")
    leaf_1.resizable(False, False)  # gWin不可改变大小
    lb_1 = tk.Label(leaf_1, text="电影类型", bg="red", fg="white", font=('黑体', 20, 'bold'))
    lb_1.grid(row=0, column=0,  sticky="EW")
    lb_2 = tk.Label(leaf_1, text="{}".format(sheet['H'][num+1].value), bg="white", fg="black", font=('宋体', 20, 'bold'))
    lb_2.grid(row=1, column=0, sticky="EW")
    leaf_1.columnconfigure(0, weight=1)

def show_people():#显示演员和导演
    global leaf_4
    leaf_4 = tk.Tk()
    leaf_4.title("Python电影")
    leaf_4.geometry("520x325")
    leaf_4.resizable(False, False)  # gWin不可改变大小
    lb_1 = tk.Label(leaf_4, text="电影导演与主演", bg="red", fg="white", font=('黑体', 20, 'bold'))
    lb_1.grid(row=0, column=0,  sticky="EW")
    actor_list = sheet['C'][num+1].value.split('\xa0\xa0\xa0')
    lb_2 = tk.Label(leaf_4, text="{}".format(actor_list[0]), bg="white", fg="black", font=('宋体', 13, 'bold'))
    lb_2.grid(row=1, column=0, sticky="EW")
    lb_2 = tk.Label(leaf_4, text="{}".format(actor_list[1]), bg="white", fg="black", font=('宋体', 13, 'bold'))
    lb_2.grid(row=2, column=0, sticky="EW")
    leaf_4.columnconfigure(0, weight=1)

def show_place():#显示地点
    global leaf_3
    leaf_3 = tk.Tk()
    leaf_3.title("Python电影")
    leaf_3.geometry("520x325")
    leaf_3.resizable(False, False)  # gWin不可改变大小
    lb_1 = tk.Label(leaf_3, text="电影产地", bg="red", fg="white", font=('黑体', 20, 'bold'))
    lb_1.grid(row=0, column=0,  sticky="EW")

    lb_2 = tk.Label(leaf_3, text="{}".format(sheet['G'][num+1].value), bg="white", fg="black", font=('宋体', 20, 'bold'))
    lb_2.grid(row=1, column=0, sticky="EW")
    leaf_3.columnconfigure(0, weight=1)

def show_year():#显示出品年
    global leaf_2
    leaf_2 = tk.Tk()
    leaf_2.title("Python电影")
    leaf_2.geometry("520x325")
    leaf_2.resizable(False, False)
    lb_1 = tk.Label(leaf_2, text="电影出版年份", bg="red", fg="white", font=('黑体', 20, 'bold'))
    lb_1.grid(row=0, column=0,  sticky="EW")
    lb_2 = tk.Label(leaf_2, text="{}".format(sheet['F'][num+1].value), bg="white", fg="black", font=('宋体', 20, 'bold'))

    lb_2.grid(row=1, column=0, sticky="EW")
    leaf_2.columnconfigure(0, weight=1)
    # leaf_2.mainloop()

def show_all():#显示全部信息
    global leaf_5
    leaf_5 = tk.Tk()
    leaf_5.title("Python电影")
    leaf_5.geometry("520x325")
    leaf_5.resizable(False, False)
    lb_1 = tk.Label(leaf_5, text="电影全部信息", bg="red", fg="white", font=('黑体', 20, 'bold'))
    lb_1.grid(row=0, column=0,  sticky="EW")
    leaf_5.nameList = tk.Listbox(leaf_5, selectmode=tk.SINGLE, exportselection=False)
    leaf_5.nameList.grid(row=1, column=0, padx=3, pady=3, sticky="NSWE")

    all_info_list = []
    all_info_list.append("年份："+sheet['F'][num+1].value)
    all_info_list.append("类型：" + sheet['H'][num + 1].value)
    all_info_list.append("产地：" + sheet['G'][num + 1].value)
    actor_list = sheet['C'][num + 1].value.split('\xa0\xa0\xa0')
    all_info_list.append(actor_list[0])
    all_info_list.append(actor_list[1])
    all_info_list.append("推荐语：" + sheet['D'][num + 1].value)
    all_info_list.append("网址：" + sheet['E'][num + 1].value)
    for x in all_info_list:
        leaf_5.nameList.insert(tk.END, x)
    leaf_5.columnconfigure(0, weight=1)
    leaf_5.rowconfigure(0, weight=1)

def grade_record():#显示用户评分记录
    global leaf_6
    leaf_6 = tk.Tk()
    leaf_6.title("Python电影")
    leaf_6.geometry("520x325")
    leaf_6.resizable(False, False)
    lb_1 = tk.Label(leaf_6, text="用户评分记录", bg="red", fg="white", font=('黑体', 20, 'bold'))
    lb_1.grid(row=0, column=0,columnspan=2, sticky="EW")
    lb_2 = tk.Label(leaf_6, text="{}".format(username), bg="white", fg="black",
                    font=('宋体', 10, 'bold'))
    lb_2.grid(row=1, column=0, sticky="W")
    bt_clear = tk.Button(leaf_6,text="清除评分记录",command = clear_grade_record)
    bt_clear.grid(row = 1,column = 1,padx=5, pady=5)
    leaf_6.gradeList = tk.Listbox(leaf_6, selectmode=tk.SINGLE, exportselection=False)
    leaf_6.gradeList.grid(row=2, column=0, columnspan=2, padx=3, pady=3, sticky="NSWE")
    for i in range(1,len(ls_1)):
        leaf_6.gradeList.insert(tk.END, ls_1[i])

    leaf_6.columnconfigure(0, weight=1)
    leaf_6.rowconfigure(0, weight=1)

def clear_grade_record():#清除用户的评分记录
    global leaf_6, username_row

    if username_row!=-1: #username_row!=-1 说明这个用户名在以前出现过，那么先把以前的记录删除
        sheet_1.delete_rows(username_row+1)
    for i in range(1,len(ls_1)):
        leaf_6.gradeList.delete(0)#删除列表框中的记录
    username_row = -1
    ls_1.clear()#删除当前的评分记录，比如这个用户是新用户，但是这次登陆进行了评分，那么也要删除
    ls_1.append(username)

def return_show_list():#返回核心界面，下同
    root1.destroy()
    show_movie_list()

def return_show_list_1():
    root2.destroy()
    show_movie_list()

def return_show_list_2():
    root3.destroy()
    show_movie_list()

def show_intro():#显示电影评分系统的使用说明
    gWin.destroy()
    global root3
    root3=tk.Tk()
    root3.title("Python电影")
    root3.geometry("520x325")
    root3.resizable(False, False)  # gWin不可改变大小
    lb_1 = tk.Label(root3, text="Python电影评分系统使用说明", bg="red", fg="white", font=('黑体', 20, 'bold'))
    lb_1.grid(row=0, column=0, columnspan=2, sticky="EW")
    btBack = tk.Button(root3, text="返回", command=return_show_list_2)
    btBack.grid(row=1, column=1, sticky="W")
    lb_text = tk.Label(root3, text="python电影评分系统中存有250部电影。\n\n用户可以选中某部电影，点击查询按键，\n\n分类查询电影的相关信息。点击查询，\n\n"+
    "还可以查询该用户的评分记录。点击评分按键，\n\n可以对选中的电影进行0-10分内的评分，\n\n按清空键可以清除用户的评分记录",
                       bg="white", fg="black",
                       font=('宋体', 15, 'bold'))
    lb_text.grid(row=2, column=0,columnspan = 2, sticky="NSWE")
    root3.columnconfigure(0, weight=1)
    root3.rowconfigure(2, weight=1)

def main():#以登录界面为主窗口构建的主函数
    gSign_in = tk.Tk()  # 创建窗口
    gSign_in.title("Python电影评分系统")
    gSign_in.geometry("500x350")
    gSign_in.resizable(False, False)
    tk.Button(gSign_in, text="登录", command=passwordDialog,width = 10).grid(row=0, column=0)
    gSign_in.pilImage = Image.open("屏幕截图 2021-08-29 122644.png")#加载一个登录界面中的图片
    gSign_in.tkImage = ImageTk.PhotoImage(image=gSign_in.pilImage)
    gSign_in.lb_3 = tk.Label(gSign_in,image = gSign_in.tkImage,width =313,height = 313)
    gSign_in.lb_3.grid(row=1, column=0,sticky="EW")
    gSign_in.columnconfigure(0, weight=1)
    gSign_in.rowconfigure(0, weight=1)
    gSign_in.mainloop()  # 显示窗口
    sheet_1.append(ls_1)
    book_1.save("用户评分记录.xlsx")

main()